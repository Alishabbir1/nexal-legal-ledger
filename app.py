"""
Nexal Legal — Client Ledger & Compliance System
Flask-based web interface for browser access
"""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from decimal import Decimal
from datetime import datetime
import sys
import os
import io
import csv
import logging
from typing import Tuple, Optional, Dict

from lib.portal_auth import (
    get_portal_dashboard_url,
    get_portal_login_url,
    get_portal_users_url,
    portal_login_redirect,
    portal_logout_redirect,
)
from lib.subscription_packages import DEFAULT_TIER, max_users_for_tier, package_display_label
from database import Database
from date_utils import (
    log_transaction_date_saved,
    month_outside_current_warning,
    parse_transaction_date_strict,
)

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


def _validated_transaction_date(form_key: str = 'transaction_date', context: str = '') -> str:
    """Parse POST date as strict YYYY-MM-DD, log it, flash if outside current calendar month."""
    iso = parse_transaction_date_strict(request.form.get(form_key), form_key)
    log_transaction_date_saved(iso, context)
    w = month_outside_current_warning(iso)
    if w:
        flash(w, 'warning')
    return iso


def _get_bundle_dir() -> str:
    """Return the directory where bundled data files (templates, static) live."""
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


def _get_data_dir() -> str:
    """Return the writable user-data directory for database, logs, backups."""
    if getattr(sys, 'frozen', False):
        d = os.path.join(
            os.environ.get('LOCALAPPDATA', os.path.expanduser('~')),
            'SolicitorLedger'
        )
    else:
        from nexal_platform.config import get_runtime_data_root

        d = get_runtime_data_root()
    os.makedirs(d, exist_ok=True)
    return d


_bundle_dir = _get_bundle_dir()
_template_dir = os.path.join(_bundle_dir, 'templates')
_static_dir = os.path.join(_bundle_dir, 'static')
_app = Flask(__name__, template_folder=_template_dir, static_folder=_static_dir)

from nexal_platform.production_secrets import DEV_FLASK_SECRET, is_production_deploy, validate_production_secrets
from nexal_platform.ops_secret import bootstrap_ledger_env, get_expected_ops_secret, get_flask_secret

bootstrap_ledger_env()

_flask_secret = get_flask_secret()
if not _flask_secret and not is_production_deploy():
    _flask_secret = DEV_FLASK_SECRET
validate_production_secrets(
    sso_secret=os.environ.get("SSO_SECRET_KEY") or os.environ.get("NEXAL_SSO_SECRET"),
    flask_secret=_flask_secret,
    ops_secret=get_expected_ops_secret() or None,
)
app = _app
app.secret_key = _flask_secret
app.config['PERMANENT_SESSION_LIFETIME'] = 900  # 15 minutes inactivity

# Legacy default database (single-tenant / pre-SSO)
_legacy_db = Database()
_legacy_db.initialize_security_columns()


def _resolve_active_db():
    """Route database access to firm tenant DB when session is firm-scoped."""
    from flask import g, has_request_context, session
    if not has_request_context():
        return _legacy_db
    cached = getattr(g, "_nexal_active_db", None)
    if cached is not None:
        return cached
    firm_id = session.get("firm_id")
    if firm_id and session.get("user_id"):
        from db_router import get_db_for_firm
        try:
            g._nexal_active_db = get_db_for_firm(firm_id)
            return g._nexal_active_db
        except (KeyError, PermissionError, OSError):
            pass
        except Exception as exc:
            logger.exception(
                "Tenant database routing failed for firm_id=%s: %s",
                firm_id,
                exc,
            )
    return _legacy_db


class _ActiveDatabaseProxy:
    def __getattr__(self, item):
        return getattr(_resolve_active_db(), item)


db = _ActiveDatabaseProxy()

# Start in-process fallback scheduler (fires at 03:00 if Task Scheduler is absent)
try:
    from backup_scheduler import start_backup_scheduler
    start_backup_scheduler(_legacy_db)
except Exception:
    pass

# Register Windows Task Scheduler task on first startup (idempotent)
try:
    from task_scheduler import ensure_task_registered

    def _task_audit(action, details):
        try:
            _legacy_db.insert_audit_log('System', 'admin', action, 'Backup System', None, details)
        except Exception:
            pass

    ensure_task_registered(audit_callback=_task_audit)
except Exception:
    pass

from firm_middleware import register_sso_routes
register_sso_routes(app)

from nexal_platform.ops_routes import register_ops_routes
register_ops_routes(app)

LOGIN_EXEMPT_ENDPOINTS = {
    'login', 'static', 'admin_recovery', 'admin_recovery_reset', 'reset_password',
    'admin_reset_password_page', 'force_password_change', 'sso_login', 'api_sso_login',
    'sso_status', 'sso_logout', 'api_ops_backup_health',
}


def get_dashboard_alerts(session_obj, database) -> list:
    """
    Return list of alert dicts for dashboard banners.
    Each dict: {type, msg, link, link_text}
    Fails safely: returns [] on any exception.
    Admin-only; non-admins get empty list.
    """
    try:
        if session_obj.get('role') != 'admin':
            return []
        today = datetime.now().date()
        alerts = []
        # Reconciliation overdue
        last_rec = database.get_most_recent_reconciliation_date()
        if last_rec:
            try:
                if (today - datetime.strptime(str(last_rec)[:10], '%Y-%m-%d').date()).days > 30:
                    alerts.append({
                        'type': 'warning',
                        'msg': 'Reconciliation overdue (>30 days)',
                        'link': '/reconciliation',
                        'link_text': 'Review Reconciliation',
                    })
            except (ValueError, TypeError):
                pass
        # Pending cheques
        pending_old = database.get_pending_cheques_older_than_days(5)
        if pending_old:
            alerts.append({
                'type': 'info',
                'msg': f'{len(pending_old)} pending cheque(s) older than 5 days',
                'link': '/cashbook',
                'link_text': 'View Cashbook',
            })
        return alerts
    except Exception:
        return []


def current_username() -> str:
    return session.get('username', 'System')


def current_role() -> str:
    return session.get('role', '')


def is_january_2026(date_text: str) -> bool:
    """Helper for controlled January 2026 reset window."""
    return '2026-01-01' <= str(date_text) <= '2026-01-31'


def _post_transaction_validation():
    """After transaction: verify ledger consistency, log mismatch to audit."""
    try:
        err = db.verify_ledger_consistency()
        if err:
            db.insert_audit_log(session.get('username', 'System'), session.get('role', 'Unknown'),
                               'Ledger consistency warning', 'Validation', None, err)
    except Exception:
        pass


def log_audit(module: str, action: str, record_id: str = None, details: str = None,
              username: str = None, role: str = None):
    """Insert audit log. Uses session username/role if not provided."""
    u = username or session.get('username', 'System')
    r = role or session.get('role', 'Unknown')
    try:
        db.insert_audit_log(u, r, action, module, record_id, details)
    except Exception:
        pass  # Never break app flow for audit failure


CLIENT_DETAIL_FIELD_LABELS = {
    'client_name': 'Client Name',
    'address': 'Address',
    'postcode': 'Postcode',
    'telephone': 'Telephone',
    'email': 'Email',
    'contact_person': 'Contact Person',
    'matter_reference': 'Reference Information',
    'description': 'Notes',
}


def can_edit_client_details(role: str = None) -> bool:
    """Admin and client-operation roles may amend client profile fields."""
    from lib.permissions import can_edit_client_details as portal_can_edit_clients

    return portal_can_edit_clients()


def _validate_client_detail_form(form) -> Tuple[Dict, Optional[str]]:
    """Parse and validate client detail edit form. Returns (updates dict, error message)."""
    import re
    updates = {}
    for field in Database.EDITABLE_CLIENT_FIELDS:
        if field in form:
            updates[field] = form.get(field)
    name = (updates.get('client_name') or '').strip() if 'client_name' in updates else None
    if name is not None and not name:
        return {}, 'Client Name is required.'
    email = (updates.get('email') or '').strip() if updates.get('email') is not None else ''
    if email and not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        return {}, 'Please enter a valid email address.'
    phone = (updates.get('telephone') or '').strip() if updates.get('telephone') is not None else ''
    if phone and (len(phone) < 5 or len(phone) > 30):
        return {}, 'Telephone must be between 5 and 30 characters.'
    return updates, None


def require_admin(f):
    """Decorator: require admin role. Redirect staff with flash message."""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('You do not have permission — Admin access required.', 'error')
            return redirect(url_for('client_ledger'))
        return f(*args, **kwargs)
    return decorated


SESSION_TIMEOUT_SECONDS = 900  # 15 minutes

@app.before_request
def require_login():
    from time import time
    from db_router import get_db_for_firm
    from nexal_platform.session_security import clear_invalid_sso_session, validate_sso_session_binding

    endpoint = request.endpoint or ''
    if endpoint in LOGIN_EXEMPT_ENDPOINTS or endpoint.startswith('static'):
        return
    if endpoint == 'logout':
        return
    if session.get('sso_login'):
        binding_error = validate_sso_session_binding(session, get_db_for_firm)
        if binding_error:
            clear_invalid_sso_session(session)
            flash('Your session is no longer valid. Please sign in again.', 'error')
            return portal_login_redirect(next_path=request.path, reason='sso_invalid')
    if session.get('user_id') and not session.get('sso_login'):
        session.clear()
        flash('Please sign in through the Portal.', 'info')
        return portal_login_redirect(next_path=request.path, reason='sso_required')
    if not session.get('user_id'):
        next_url = request.path if request.method == 'GET' else url_for('client_ledger')
        return portal_login_redirect(next_path=next_url)
    session.permanent = True
    last = session.get('_last_activity', time())
    now = time()
    if now - last > SESSION_TIMEOUT_SECONDS:
        u, r = session.get('username', 'Unknown'), session.get('role', 'Unknown')
        session.clear()
        try:
            db.insert_audit_log(u, r, 'Logout - Session timeout', 'Authentication', None, 'Session expired after 15 min inactivity')
        except Exception:
            pass
        flash('Session expired due to inactivity. Please log in again.', 'info')
        return portal_login_redirect(next_path=request.path, reason='timeout')
    session['_last_activity'] = now


@app.context_processor
def inject_user():
    from lib.firm_package import package_usage_summary, resolve_package_display_for_request

    from lib.branding import LEGAL_COMPANY_NAME, PRODUCT_NAME

    ctx = {
        'current_username': session.get('username'),
        'current_role': session.get('role'),
        'legal_company_name': LEGAL_COMPANY_NAME,
        'product_name': PRODUCT_NAME,
        'dashboard_alerts': get_dashboard_alerts(session, db),
        'override_mode_active': session.get('override_mode', False),
        'override_mode_reason': session.get('override_reason', ''),
        'override_mode_by': session.get('override_enabled_by', ''),
        'override_mode_at': session.get('override_enabled_at', ''),
    }
    try:
        ctx['firm_package_label'] = resolve_package_display_for_request(session, db)
        ctx['firm_package_usage'] = package_usage_summary(db, session)
    except Exception:
        fallback_label = package_display_label(DEFAULT_TIER)
        ctx['firm_package_label'] = fallback_label
        ctx['firm_package_usage'] = {
            'tier': DEFAULT_TIER,
            'label': fallback_label,
            'active_users': _legacy_db.count_billable_active_users(),
            'max_users': max_users_for_tier(DEFAULT_TIER),
            'at_limit': False,
        }
    ctx['portal_dashboard_url'] = get_portal_dashboard_url()
    ctx['portal_users_url'] = get_portal_users_url()
    return ctx


@app.route('/')
def index():
    """Main dashboard"""
    return redirect(url_for('client_ledger'))


# ─── Override Premium Mode ────────────────────────────────────────────────────

@app.route('/admin/override-mode', methods=['GET', 'POST'])
@require_admin
def override_mode():
    """Override Premium Mode management page."""
    total_cashbook = db.get_total_cashbook_net_balance()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'enable':
            reason = (request.form.get('reason') or '').strip()
            if not reason:
                flash('A reason is required to enable Override Premium Mode.', 'error')
                return redirect(url_for('override_mode'))
            from datetime import datetime as _dt
            session['override_mode'] = True
            session['override_reason'] = reason
            session['override_enabled_by'] = current_username()
            session['override_enabled_at'] = _dt.now().strftime('%Y-%m-%d %H:%M:%S')
            log_audit('Override', 'OVERRIDE_MODE_ENABLED',
                      details=f"Override Premium Mode ENABLED | Reason: {reason} | "
                              f"Total cashbook at activation: £{total_cashbook:,.2f}")
            flash('Override Premium Mode activated. Transactions may now use the full client cashbook balance. '
                  'All actions are audit-logged.', 'warning')
            return redirect(url_for('override_mode'))
        elif action == 'disable':
            override_reason = session.get('override_reason', '')
            override_by = session.get('override_enabled_by', '')
            session.pop('override_mode', None)
            session.pop('override_reason', None)
            session.pop('override_enabled_by', None)
            session.pop('override_enabled_at', None)
            log_audit('Override', 'OVERRIDE_MODE_DISABLED',
                      details=f"Override Premium Mode DISABLED by {current_username()} | "
                              f"Previous reason: {override_reason} | Enabled by: {override_by}")
            flash('Override Premium Mode deactivated. Strict per-client ledger enforcement restored.', 'success')
            return redirect(url_for('override_mode'))
    return render_template('override_mode.html', total_cashbook=total_cashbook)


@app.route('/admin/override-impact')
@require_admin
def override_impact():
    """Display transaction impact summary after an override transaction."""
    impact = session.pop('override_impact', None)
    if not impact:
        return redirect(url_for('cashbook'))
    return render_template('override_impact.html', impact=impact)


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Phase 4E: Ledger is SSO-only — redirect all login attempts to the Portal."""
    if session.get('user_id') and session.get('sso_login'):
        return redirect(url_for('client_ledger'))
    session.pop('firm_id', None)
    next_path = request.args.get('next') or request.form.get('next')
    reason = request.args.get('reason') or ('direct_login_disabled' if request.method == 'POST' else None)
    return portal_login_redirect(next_path=next_path, reason=reason)


@app.route('/logout')
def logout():
    """Destroy Ledger session and return the user to the Portal."""
    username = session.get('username', 'Unknown')
    role = session.get('role', 'Unknown')
    reason = request.args.get('reason') or 'User logout'
    log_audit('Authentication', 'Logout', details=f"{username} | {reason}", username=username, role=role)
    session.clear()
    flash('You have been logged out of Ledger.', 'info')
    return portal_logout_redirect()


@app.route('/client-ledger')
def client_ledger():
    """Client ledger page"""
    clients = db.get_all_clients()
    selected_client_id = request.args.get('client_id', type=int)
    transactions = []
    current_balance = Decimal('0')
    selected_client = None
    client_office_total = Decimal('0')
    fee_transfer_ledger_ids = set()

    if selected_client_id:
        selected_client = db.get_client(selected_client_id)
        transactions = db.get_client_transactions(selected_client_id)
        current_balance = db.get_client_balance(selected_client_id)
        # Fee transfers: sum office income from this client and build id lookup
        all_fee_transfers = db.get_office_fee_transfers()
        client_fee_transfers = [
            ft for ft in all_fee_transfers
            if ft.get('client_id') == selected_client_id and not ft.get('is_deleted')
        ]
        client_office_total = sum(Decimal(str(ft['amount'])) for ft in client_fee_transfers)
        fee_transfer_ledger_ids = {
            ft['ledger_transaction_id']
            for ft in client_fee_transfers
            if ft.get('ledger_transaction_id')
        }

    return render_template('client_ledger.html',
                           clients=clients,
                           selected_client=selected_client,
                           transactions=transactions,
                           current_balance=current_balance,
                           client_office_total=client_office_total,
                           fee_transfer_ledger_ids=fee_transfer_ledger_ids)


def generate_client_code():
    """Generate next CLC-XXXX-XXXX code. Reads stored number, increments, returns new code. Ignores any browser input."""
    return db.reserve_next_client_code()


def _render_new_client_form(auto_client_code, client_name='', matter_reference='', description=''):
    """Render New Client form with optional prefilled values (for validation errors)."""
    return render_template('new_client.html',
                          auto_client_code=auto_client_code,
                          client_name=client_name,
                          matter_reference=matter_reference,
                          description=description)


@app.route('/client-ledger/new-client', methods=['GET', 'POST'])
def new_client():
    """Create new client with auto-generated CLC-XXXX-XXXX code. Client code is ALWAYS server-generated, never from browser."""
    if request.method == 'POST':
        client_name = (request.form.get('client_name') or '').strip()
        matter_reference = (request.form.get('matter_reference') or '').strip() or None
        description = (request.form.get('description') or '').strip() or None

        if not client_name:
            flash('Client Name is required.', 'error')
            code = session.get('reserved_client_code') or generate_client_code()
            if not session.get('reserved_client_code'):
                session['reserved_client_code'] = code
            return _render_new_client_form(code, client_name or '', matter_reference or '', description or '')

        reserved_code = session.pop('reserved_client_code', None)
        if not reserved_code:
            reserved_code = generate_client_code()

        try:
            client_id = db.create_client(
                client_code=reserved_code,
                client_name=client_name,
                matter_reference=matter_reference,
                description=description
            )
            client = db.get_client(client_id)
            log_audit('Client Ledger', 'Client created', record_id=str(client_id),
                      details=f"{client['client_code']} - {client['client_name']}")
            flash(f'Client created successfully. Code: {client["client_code"]}', 'success')
            return redirect(url_for('client_ledger', client_id=client_id))
        except ValueError as e:
            flash(str(e), 'error')
            session['reserved_client_code'] = reserved_code
            return _render_new_client_form(reserved_code, client_name, matter_reference or '', description or '')
        except Exception as e:
            flash(f'Error saving client: {str(e)}', 'error')
            session['reserved_client_code'] = reserved_code
            return _render_new_client_form(reserved_code, client_name, matter_reference or '', description or '')

    auto_client_code = generate_client_code()
    session['reserved_client_code'] = auto_client_code
    return _render_new_client_form(auto_client_code)


@app.route('/client-ledger/edit-client/<int:client_id>', methods=['POST'])
def edit_client_details(client_id):
    """Amend client profile fields. Financial records are not modified."""
    if not session.get('user_id'):
        return redirect(url_for('login'))
    if not can_edit_client_details():
        flash('You do not have permission to edit client details.', 'error')
        return redirect(url_for('client_ledger', client_id=client_id))

    client = db.get_client(client_id)
    if not client:
        flash('Client not found.', 'error')
        return redirect(url_for('client_ledger'))

    updates, err = _validate_client_detail_form(request.form)
    if err:
        flash(err, 'error')
        return redirect(url_for('client_ledger', client_id=client_id))

    try:
        changes, updated = db.update_client_details(
            client_id, updates, updated_by=current_username(),
        )
    except ValueError as e:
        flash(str(e), 'error')
        return redirect(url_for('client_ledger', client_id=client_id))

    if not changes:
        flash('No changes were made.', 'info')
        return redirect(url_for('client_ledger', client_id=client_id))

    detail_lines = [
        f"{CLIENT_DETAIL_FIELD_LABELS.get(c['field'], c['field'])}: "
        f"'{c['old_value'] or '—'}' → '{c['new_value'] or '—'}'"
        for c in changes
    ]
    log_audit(
        'Client Ledger',
        'Client Details Updated',
        record_id=str(client_id),
        details=(
            f"{updated['client_code']} — {updated['client_name']}\n"
            + '\n'.join(detail_lines)
        ),
    )
    flash('Client details updated successfully.', 'success')
    return redirect(url_for('client_ledger', client_id=client_id))


@app.route('/client-ledger/new-transaction', methods=['GET', 'POST'])
def new_transaction():
    """Create new ledger transaction"""
    clients = db.get_all_clients()
    client_id = request.args.get('client_id', type=int) or request.form.get('client_id', type=int)
    override_active = session.get('override_mode', False)

    if request.method == 'POST':
        try:
            cid = int(request.form['client_id'])
            source = request.form['source']
            txn_type = request.form['transaction_type']
            amount = Decimal(request.form['amount'])
            reference = request.form['reference']
            transaction_date = _validated_transaction_date('transaction_date', 'client ledger + cashbook')

            # Capture pre-transaction balances
            pre_client_balance = db.get_cleared_client_balance(cid)
            pre_total_cashbook = db.get_total_cashbook_net_balance()

            lid, cbid, txn_id = db.create_ledger_and_cashbook_transaction(
                client_id=cid,
                transaction_date=transaction_date,
                amount=amount,
                transaction_type=txn_type,
                reference=reference,
                source=source,
                description=request.form.get('description') or None,
                created_by=current_username(),
                allow_override=override_active
            )

            # Capture post-transaction balances
            post_client_balance = db.get_cleared_client_balance(cid)
            post_total_cashbook = db.get_total_cashbook_net_balance()
            client_deficit = post_client_balance < Decimal('0')
            override_used = override_active and txn_type in ('Payment', 'Transfer') and pre_client_balance < amount

            if source == 'Cheque':
                log_audit('Client Ledger', 'CHEQUE_CREATED', record_id=str(lid),
                          details=f"TXN {txn_id} | Pending cheque {txn_type} £{amount} ref {reference}. Balance not affected until cleared.")
            else:
                log_audit('Client Ledger', 'Transaction created', record_id=str(lid),
                          details=f"TXN {txn_id} | {txn_type} £{amount} ref {reference}")

            if override_active:
                selected_client_obj = db.get_client(cid)
                log_audit('Override', 'OVERRIDE_TRANSACTION_COMPLETED', record_id=str(lid),
                          details=(
                              f"TXN {txn_id} | Override {txn_type} £{amount:,.2f} | "
                              f"Client ledger: £{pre_client_balance:,.2f} → £{post_client_balance:,.2f} | "
                              f"Total cashbook: £{pre_total_cashbook:,.2f} → £{post_total_cashbook:,.2f} | "
                              f"{'DEFICIT CREATED' if client_deficit else 'No deficit'} | "
                              f"Reason: {session.get('override_reason', '')}"
                          ))
                session['override_impact'] = {
                    'amount': str(amount),
                    'txn_type': txn_type,
                    'reference': reference,
                    'source': source,
                    'client_id': cid,
                    'client_name': selected_client_obj.get('client_name', '') if selected_client_obj else '',
                    'matter_ref': selected_client_obj.get('matter_reference', '') if selected_client_obj else '',
                    'pre_client_balance': str(pre_client_balance),
                    'post_client_balance': str(post_client_balance),
                    'pre_total_cashbook': str(pre_total_cashbook),
                    'post_total_cashbook': str(post_total_cashbook),
                    'override_used': override_used,
                    'client_deficit': str(post_client_balance) if client_deficit else None,
                    'override_reason': session.get('override_reason', ''),
                    'override_by': session.get('override_enabled_by', ''),
                    'override_at': session.get('override_enabled_at', ''),
                    'transaction_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'ledger_id': lid,
                    'txn_id': txn_id,
                }
                _post_transaction_validation()
                return redirect(url_for('override_impact'))

            if source == 'Cheque':
                flash(f'Cheque {txn_type.lower()} created as PENDING. Client balance will update when the cheque is cleared.', 'info')
            else:
                flash('Transaction synced to Cashbook successfully.', 'success')
            _post_transaction_validation()
            return redirect(url_for('cashbook'))
        except ValueError as e:
            if 'Insufficient cleared client funds' in str(e):
                log_audit('Client Ledger', 'Transaction blocked - insufficient funds', details=str(e))
            elif 'Override blocked' in str(e):
                log_audit('Override', 'OVERRIDE_TRANSACTION_BLOCKED', details=str(e))
            flash(str(e), 'error')

    current_balance = Decimal('0')
    cleared_balance = Decimal('0')
    total_cashbook = Decimal('0')
    selected_client = None
    if client_id:
        current_balance = db.get_client_balance(client_id)
        cleared_balance = db.get_cleared_client_balance(client_id)
        selected_client = db.get_client(client_id)
    total_cashbook = db.get_total_cashbook_net_balance()

    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('new_transaction.html',
                         clients=clients,
                         selected_client_id=client_id,
                         selected_client=selected_client,
                         current_balance=current_balance,
                         cleared_balance=cleared_balance,
                         total_cashbook=total_cashbook,
                         override_active=override_active,
                         today=today)


@app.route('/client-ledger/transfer-fee-to-office', methods=['GET', 'POST'])
def transfer_fee_to_office():
    """Transfer earned fee to office: ledger Payment + linked cashbook Payment + office_fee_transfer (atomic)."""
    clients = db.get_all_clients()
    client_id = request.args.get('client_id', type=int) or (request.form.get('client_id', type=int) if request.method == 'POST' else None)
    
    if request.method == 'POST':
        try:
            client_id = int(request.form['client_id'])
            amount = Decimal(request.form['amount'])
            reference = request.form['reference']
            transaction_date = request.form['transaction_date']
            description = request.form.get('description') or None
            
            result = db.create_transfer_fee_to_office(
                client_id=client_id,
                transaction_date=transaction_date,
                amount=amount,
                reference=reference,
                description=description,
                created_by=current_username(),
            )
            matter = result.get('matter_reference') or '—'
            log_audit(
                'Office Account',
                'Office Transfer',
                record_id=str(result['ledger_id']),
                details=(
                    f"£{amount:,.2f} | Client {result['client_code']} | Matter {matter} | "
                    f"Ledger #{result['ledger_id']} | Cashbook #{result['cashbook_id']} | "
                    f"User {current_username()}"
                ),
            )
            _post_transaction_validation()
            flash(f'Fee of £{amount:,.2f} transferred to Office Account successfully', 'success')
            return redirect(url_for('client_ledger', client_id=client_id))
        except ValueError as e:
            flash(str(e), 'error')
    
    current_balance = Decimal('0')
    cleared_balance = Decimal('0')
    selected_client = None
    if client_id:
        current_balance = db.get_client_balance(client_id)
        cleared_balance = db.get_cleared_client_balance(client_id)
        selected_client = db.get_client(client_id)
    
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('transfer_fee_to_office.html',
                         clients=clients,
                         selected_client_id=client_id,
                         selected_client=selected_client,
                         current_balance=current_balance,
                         cleared_balance=cleared_balance,
                         today=today)


@app.route('/client-ledger/close-matter/<int:client_id>', methods=['POST'])
@require_admin
def close_matter(client_id):
    """
    Close matter - blocks new transactions until reopened.
    
    SRA Rule: A matter may only be closed when client_balance == 0.
    This prevents hidden client funds, reconciliation inconsistencies, and audit failures.
    """
    # Check client balance before allowing close
    balance = db.get_client_balance(client_id)
    if balance != Decimal('0'):
        flash(
            f'Matter cannot be closed while client funds remain (current balance: £{balance:,.2f}). '
            'Please transfer or return all funds before closing.',
            'error'
        )
        log_audit('Client Ledger', 'MATTER_CLOSE_BLOCKED', record_id=str(client_id),
                 details=f"Close blocked - balance £{balance:,.2f} exists")
        return redirect(url_for('client_ledger', client_id=client_id))
    
    try:
        db.set_matter_status(client_id, 'CLOSED', current_username())
        log_audit('Client Ledger', 'MATTER_CLOSED', record_id=str(client_id),
                 details=f"Client {client_id} closed with £0 balance")
        flash('Matter closed. No new transactions can be added until reopened.', 'success')
    except ValueError as e:
        flash(str(e), 'error')
    return redirect(url_for('client_ledger', client_id=client_id))


@app.route('/client-ledger/reopen-matter/<int:client_id>', methods=['POST'])
@require_admin
def reopen_matter(client_id):
    """Reopen closed matter."""
    try:
        db.set_matter_status(client_id, 'OPEN', current_username())
        log_audit('Client Ledger', 'Matter reopened', record_id=str(client_id), details=f"Client {client_id}")
        flash('Matter reopened. Transactions can be added again.', 'success')
    except ValueError as e:
        flash(str(e), 'error')
    return redirect(url_for('client_ledger', client_id=client_id))


@app.route('/client-ledger/reverse/<int:ledger_id>', methods=['POST'])
@require_admin
def reverse_ledger_entry(ledger_id):
    """Reverse a ledger transaction (admin only). Creates compensating entry, marks original REVERSED."""
    client_id = request.form.get('client_id', type=int)
    reason = (request.form.get('reversal_reason') or '').strip()
    print(f"POST reverse_ledger_entry ledger_id={ledger_id} client_id={client_id} user={session.get('username')!r}")
    if not reason:
        flash('Reversal reason is mandatory.', 'error')
        return redirect(url_for('client_ledger', client_id=client_id))
    try:
        rev_raw = (request.form.get('reversal_date') or '').strip()
        if rev_raw:
            reversal_transaction_date = parse_transaction_date_strict(rev_raw, 'reversal_date')
            log_transaction_date_saved(reversal_transaction_date, 'ledger reversal')
            mw = month_outside_current_warning(reversal_transaction_date)
            if mw:
                flash(mw, 'warning')
        else:
            reversal_transaction_date = datetime.now().strftime('%Y-%m-%d')
        result = db.reverse_ledger_transaction(
            ledger_id,
            current_username(),
            reason,
            reversal_transaction_date=reversal_transaction_date,
        )
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        new_id = result['reversal_ledger_id']
        print(f"Reversal committed: original_transaction_id={ledger_id} new_transaction_id={new_id}")
        log_audit(
            'Client Ledger',
            'REVERSAL',
            record_id=str(new_id),
            details=(
                f"original_transaction_id={ledger_id} new_transaction_id={new_id} "
                f"user={current_username()} timestamp={ts} reason={reason!r} | "
                f"reversal_txn_ref={result['reversal_ledger_txn_id']} | "
                f"type {result['original_type']} -> {result['reversal_type']} amount £{result['amount']} "
                f"client {result['client_id']} reversal_date={reversal_transaction_date}"
            ),
        )
        _post_transaction_validation()
        flash(f"Transaction reversed successfully. Reversal ID: {result['reversal_ledger_txn_id']}", 'success')
        if result.get('mismatch_client') or result.get('mismatch_total'):
            log_audit(
                'Client Ledger',
                'REVERSAL_POST_MISMATCH_WARNING',
                record_id=str(result.get('reversal_ledger_id')),
                details=(
                    f"client_mismatch={result.get('mismatch_client')} total_mismatch={result.get('mismatch_total')} "
                    f"post_ledger={result.get('post_ledger')} post_cashbook={result.get('post_cashbook')} "
                    f"total_ledger={result.get('total_ledger')} total_cashbook={result.get('total_cashbook')}"
                ),
            )
            flash(
                'Reversal posted. Ledger and cashbook totals differ — please reconcile if balances differ.',
                'warning',
            )
    except ValueError as e:
        logger.warning('Ledger reversal failed: %s', e)
        flash(str(e), 'error')
    except Exception as e:
        logger.exception('Ledger reversal unexpected error')
        flash(f'Unable to reverse transaction: {e}', 'error')
    return redirect(url_for('client_ledger', client_id=client_id))


@app.route('/cashbook')
def cashbook():
    """Client Money Cashbook - client-linked transactions only. Office transactions are in Office Account."""
    status_filter = request.args.get('status', 'All')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    transactions = db.get_all_cashbook_transactions(
        start_date=date_from if date_from else None,
        end_date=date_to if date_to else None,
        status=None if status_filter == 'All' else status_filter,
        client_only=True
    )
    
    bank_balance = db.get_bank_balance(client_only=True)
    
    # Calculate pending amount (client cashbook only)
    all_transactions = db.get_all_cashbook_transactions(client_only=True)
    pending_total = Decimal('0')
    for trans in all_transactions:
        if trans['status'] == 'Pending':
            if trans['transaction_type'] == 'Receipt':
                pending_total += Decimal(str(trans['amount']))
            else:
                pending_total -= Decimal(str(trans['amount']))
    
    return render_template('cashbook.html',
                         transactions=transactions,
                         bank_balance=bank_balance,
                         pending_total=pending_total,
                         status_filter=status_filter,
                         date_from=date_from,
                         date_to=date_to)


@app.route('/cashbook/new-transaction', methods=['GET', 'POST'])
def new_cashbook_transaction():
    """Create new cashbook transaction. Optionally link to client (creates ledger + cashbook)."""
    clients = db.get_all_clients()
    override_active = session.get('override_mode', False)

    if request.method == 'POST':
        try:
            client_id = request.form.get('client_id', type=int)
            transaction_date = _validated_transaction_date('transaction_date', 'cashbook + ledger')
            amount = Decimal(request.form['amount'])
            transaction_type = request.form['transaction_type']
            reference = request.form['reference']
            source = request.form['source']
            description = request.form.get('description') or None

            if not client_id:
                flash('For office transactions, use Add Office Income or Add Office Expense from the Office Account page.', 'info')
                return redirect(url_for('office_account'))

            # Capture pre-transaction balances
            pre_client_balance = db.get_cleared_client_balance(client_id)
            pre_total_cashbook = db.get_total_cashbook_net_balance()

            # Client-linked: create ledger transaction first, then cashbook linked to it
            ledger_id = db.create_ledger_transaction(
                client_id=client_id,
                transaction_date=transaction_date,
                amount=amount,
                transaction_type=transaction_type,
                reference=reference,
                source=source,
                description=description,
                linked_cashbook_id=None,
                created_by=current_username(),
                allow_override=override_active
            )
            cashbook_id = db.create_cashbook_transaction(
                transaction_date=transaction_date,
                amount=amount,
                transaction_type=transaction_type,
                reference=reference,
                source=source,
                description=description,
                linked_ledger_id=ledger_id,
                created_by=current_username()
            )
            db.update_ledger_linked_cashbook(ledger_id, cashbook_id)
            log_audit('Client Ledger', 'Transaction created', record_id=str(ledger_id),
                      details=f"{transaction_type} £{amount} ref {reference}")

            # Capture post-transaction balances
            post_client_balance = db.get_cleared_client_balance(client_id)
            post_total_cashbook = db.get_total_cashbook_net_balance()
            client_deficit = post_client_balance < Decimal('0')
            override_used = override_active and transaction_type in ('Payment', 'Transfer') and pre_client_balance < amount

            if override_active:
                selected_client_obj = db.get_client(client_id)
                log_audit('Override', 'OVERRIDE_TRANSACTION_COMPLETED', record_id=str(ledger_id),
                          details=(
                              f"Override {transaction_type} £{amount:,.2f} | "
                              f"Client ledger: £{pre_client_balance:,.2f} → £{post_client_balance:,.2f} | "
                              f"Total cashbook: £{pre_total_cashbook:,.2f} → £{post_total_cashbook:,.2f} | "
                              f"{'DEFICIT CREATED' if client_deficit else 'No deficit'} | "
                              f"Reason: {session.get('override_reason', '')}"
                          ))
                session['override_impact'] = {
                    'amount': str(amount),
                    'txn_type': transaction_type,
                    'reference': reference,
                    'source': source,
                    'client_id': client_id,
                    'client_name': selected_client_obj.get('client_name', '') if selected_client_obj else '',
                    'matter_ref': selected_client_obj.get('matter_reference', '') if selected_client_obj else '',
                    'pre_client_balance': str(pre_client_balance),
                    'post_client_balance': str(post_client_balance),
                    'pre_total_cashbook': str(pre_total_cashbook),
                    'post_total_cashbook': str(post_total_cashbook),
                    'override_used': override_used,
                    'client_deficit': str(post_client_balance) if client_deficit else None,
                    'override_reason': session.get('override_reason', ''),
                    'override_by': session.get('override_enabled_by', ''),
                    'override_at': session.get('override_enabled_at', ''),
                    'transaction_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'ledger_id': ledger_id,
                }
                _post_transaction_validation()
                return redirect(url_for('override_impact'))

            _post_transaction_validation()
            status_msg = 'Pending (cheque)' if source == 'Cheque' else 'Cleared'
            flash(f'Transaction created successfully. Status: {status_msg}', 'success')
            return redirect(url_for('cashbook'))
        except ValueError as e:
            if 'Insufficient cleared client funds' in str(e):
                log_audit('Client Ledger', 'Transaction blocked - insufficient funds', details=str(e))
            elif 'Override blocked' in str(e):
                log_audit('Override', 'OVERRIDE_TRANSACTION_BLOCKED', details=str(e))
            elif 'Office account balance cannot go below' in str(e):
                log_audit('Office Account', 'Transaction blocked - office balance', details=str(e))
            flash(str(e), 'error')

    total_cashbook = db.get_total_cashbook_net_balance()
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('new_cashbook_transaction.html',
                           clients=clients,
                           today=today,
                           total_cashbook=total_cashbook,
                           override_active=override_active)


@app.route('/cashbook/update-status/<int:transaction_id>', methods=['POST'])
def update_cashbook_status(transaction_id):
    """
    Update cashbook transaction status (cheque clearance/decline).
    
    Permissions:
    - Clear cheque: Admin or Staff (operational cashier function)
    - Decline cheque: Admin or Staff
    """
    try:
        new_status = request.form['status']
        reason = (request.form.get('decline_reason') or request.form.get('reason') or '').strip()
        if new_status == 'Declined' and not reason:
            flash('Decline reason is mandatory. Please provide a reason.', 'error')
            return redirect(url_for('cashbook'))
        # Cheque clearance is operational - allowed for Admin and Staff
        if new_status == 'Cleared' and current_role() not in ('admin', 'staff'):
            flash('You do not have permission — Staff or Admin access required.', 'error')
            return redirect(url_for('cashbook'))
        db.update_cashbook_status(transaction_id, new_status, reason, current_username())
        if new_status == 'Cleared':
            log_audit('Cashbook', 'CHEQUE_CLEARED', record_id=str(transaction_id), 
                      details=f"Cheque cleared by {current_username()} (role: {current_role()}). Balance now reflects this transaction.")
        else:
            log_audit('Cashbook', 'CHEQUE_DECLINED', record_id=str(transaction_id), 
                      details=f"Cheque declined by {current_username()}. Reason: {reason}")
        if new_status == 'Declined':
            flash('Cheque declined. No balance impact as cheque was pending and never affected client funds.', 'info')
        else:
            flash('Cheque cleared successfully. Client ledger balance has been updated.', 'success')
    except Exception as e:
        flash(f'Error updating status: {str(e)}', 'error')
    
    return redirect(url_for('cashbook'))


@app.route('/office-account')
def office_account():
    """Office Account page - firm's own money (profit)"""
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    transactions = db.get_office_transactions(
        start_date=date_from if date_from else None,
        end_date=date_to if date_to else None
    )
    
    office_balance = db.get_office_balance()
    office_income = db.get_office_income_total(date_from or None, date_to or None)
    office_expenses = db.get_office_expenses_total(date_from or None, date_to or None)
    
    # Calculate pending cheque total
    pending_total = Decimal('0')
    for trans in transactions:
        if trans.get('status') == 'Pending' and trans.get('office_cashbook_id'):
            if trans.get('transaction_type') == 'Receipt':
                pending_total += Decimal(str(trans.get('amount', 0)))
            else:
                pending_total -= Decimal(str(abs(trans.get('amount', 0))))
    
    return render_template('office_account.html',
                         transactions=transactions,
                         office_balance=office_balance,
                         office_income=office_income,
                         office_expenses=office_expenses,
                         pending_total=pending_total,
                         date_from=date_from,
                         date_to=date_to)


@app.route('/office-account/add-income', methods=['GET', 'POST'])
def office_add_income():
    """Add office income - writes to office_cashbook only (never client ledger/cashbook)."""
    if request.method == 'POST':
        try:
            source = request.form['source']
            amount = Decimal(request.form['amount'])
            transaction_date = _validated_transaction_date('transaction_date', 'office income')
            row_id = db.create_office_transaction(
                transaction_date=transaction_date,
                amount=amount,
                transaction_type='Receipt',
                reference=request.form['reference'],
                source=source,
                description=request.form.get('description') or None,
                created_by=current_username()
            )
            if source == 'Cheque':
                log_audit('Office Account', 'OFFICE_CHEQUE_CREATED', record_id=str(row_id),
                          details=f"Pending cheque income £{amount} ref {request.form['reference']}. Balance not affected until cleared.")
                flash(f'Office cheque income created as PENDING. Balance will update when the cheque is cleared.', 'info')
            else:
                log_audit('Office Account', 'OFFICE_RECEIPT_CREATED', record_id=str(row_id),
                          details=f"£{amount} {request.form['reference']}")
                flash('Office income recorded successfully', 'success')
            return redirect(url_for('office_account'))
        except ValueError as e:
            flash(str(e), 'error')
    
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('new_office_income.html', today=today)


@app.route('/office-account/add-expense', methods=['GET', 'POST'])
def office_add_expense():
    """Add office expense - writes to office_cashbook only (never client ledger/cashbook)."""
    if request.method == 'POST':
        try:
            source = request.form['source']
            amount = Decimal(request.form['amount'])
            transaction_date = _validated_transaction_date('transaction_date', 'office expense')
            row_id = db.create_office_transaction(
                transaction_date=transaction_date,
                amount=amount,
                transaction_type='Payment',
                reference=request.form['reference'],
                source=source,
                description=request.form.get('description') or None,
                created_by=current_username()
            )
            if source == 'Cheque':
                log_audit('Office Account', 'OFFICE_CHEQUE_CREATED', record_id=str(row_id),
                          details=f"Pending cheque expense £{amount} ref {request.form['reference']}. Balance not affected until cleared.")
                flash(f'Office cheque expense created as PENDING. Balance will update when the cheque is cleared.', 'info')
            else:
                log_audit('Office Account', 'Office expense created', record_id=str(row_id),
                          details=f"£{amount} {request.form['reference']}")
                flash('Office expense recorded successfully', 'success')
            _post_transaction_validation()
            return redirect(url_for('office_account'))
        except ValueError as e:
            if 'Office account balance cannot go below' in str(e):
                log_audit('Office Account', 'Transaction blocked - office balance', details=str(e))
            flash(str(e), 'error')
    
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('new_office_expense.html', today=today)


@app.route('/office-account/update-status/<int:transaction_id>', methods=['POST'])
def update_office_cashbook_status(transaction_id):
    """
    Update office cashbook cheque status (clearance/decline).
    
    Permissions:
    - Clear cheque: Admin or Staff (operational cashier function)
    - Decline cheque: Admin or Staff
    """
    try:
        new_status = request.form['status']
        reason = (request.form.get('decline_reason') or request.form.get('reason') or '').strip()
        if new_status == 'Declined' and not reason:
            flash('Decline reason is mandatory. Please provide a reason.', 'error')
            return redirect(url_for('office_account'))
        # Cheque clearance is operational - allowed for Admin and Staff
        if new_status == 'Cleared' and current_role() not in ('admin', 'staff'):
            flash('You do not have permission — Staff or Admin access required.', 'error')
            return redirect(url_for('office_account'))
        db.update_office_cashbook_status(transaction_id, new_status, reason, current_username())
        if new_status == 'Cleared':
            log_audit('Office Account', 'OFFICE_CHEQUE_CLEARED', record_id=str(transaction_id),
                      details=f"Office cheque cleared by {current_username()} (role: {current_role()}). Balance now reflects this transaction.")
            flash('Office cheque cleared successfully. Balance has been updated.', 'success')
        else:
            log_audit('Office Account', 'OFFICE_CHEQUE_DECLINED', record_id=str(transaction_id),
                      details=f"Office cheque declined by {current_username()}. Reason: {reason}")
            flash('Office cheque declined. No balance impact as cheque was pending.', 'info')
    except Exception as e:
        flash(f'Error updating status: {str(e)}', 'error')
    
    return redirect(url_for('office_account'))


@app.route('/reconciliation')
def reconciliation():
    """Reconciliation page"""
    locked_months = db.get_locked_months()
    current_reconciliations = db.get_current_reconciliations()
    past_reconciliations = db.get_all_reconciliations()
    return render_template('reconciliation.html',
                         locked_months=locked_months,
                         current_year=datetime.now().year,
                         current_reconciliations=current_reconciliations,
                         past_reconciliations=past_reconciliations)


@app.route('/reconciliation/lock-month', methods=['POST'])
@require_admin
def lock_month():
    """Lock reconciliation using live month-end figures; relock creates next version."""
    import calendar
    try:
        month = int(request.form['lock_month'])
        year = int(request.form['lock_year'])
        last_day = calendar.monthrange(year, month)[1]
        rec_date = f'{year}-{month:02d}-{last_day:02d}'
        ledger_total = db.get_total_ledger_balance(as_of_date=rec_date)
        cashbook_total = _client_cashbook_total_to_date(rec_date)
        bank_balance = db.get_bank_balance(as_of_date=rec_date, client_only=True)
        result = db.lock_reconciliation_month(
            month, year, current_username(), rec_date,
            ledger_total, cashbook_total, bank_balance,
        )
        period = f'{year}-{month:02d}'
        ver = result['version']
        if result['action'] == 'new_version':
            log_audit(
                'Reconciliation', 'New Version Created',
                record_id=str(result['id']),
                details=f'{period} v{ver} | ledger={ledger_total} cashbook={cashbook_total} bank={bank_balance}',
            )
            flash(f'{period} locked — version {ver} created (figures changed).', 'success')
        elif result['action'] == 'relocked_unchanged':
            log_audit(
                'Reconciliation', 'Reconciliation Locked',
                record_id=str(result['id']),
                details=f'{period} v{ver} | unchanged | ledger={ledger_total} cashbook={cashbook_total} bank={bank_balance}',
            )
            flash(f'{period} relocked — version {ver} unchanged.', 'success')
        else:
            log_audit(
                'Reconciliation', 'Reconciliation Locked',
                record_id=str(result['id']),
                details=f'{period} v{ver} | ledger={ledger_total} cashbook={cashbook_total} bank={bank_balance}',
            )
            flash(f'{period} locked successfully (version {ver}).', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'error')
    return redirect(url_for('reconciliation'))


@app.route('/reconciliation/unlock-month', methods=['POST'])
@require_admin
def unlock_month():
    """Unlock month for corrections; current version returns to live calculation mode."""
    try:
        month = int(request.form['lock_month'])
        year = int(request.form['lock_year'])
        reason = (request.form.get('unlock_reason') or '').strip() or 'No reason provided'
        result = db.unlock_reconciliation_month(month, year, current_username())
        period = f'{year}-{month:02d}'
        log_audit(
            'Reconciliation', 'Reconciliation Unlocked',
            record_id=str(result['id']),
            details=f'{period} v{result["version"]} | Reason: {reason}',
        )
        flash(
            f'{period} unlocked (v{result["version"]}). Figures will use live month-end calculations until relocked.',
            'success',
        )
    except Exception as e:
        flash(f'Error: {e}', 'error')
    return redirect(url_for('reconciliation'))


@app.route('/reconciliation/upload-bank-statement', methods=['POST'])
def upload_bank_statement():
    """Bank statement import disabled — not required for launch reconciliation workflow."""
    return redirect(url_for('reconciliation'))


@app.route('/reconciliation/run-matching', methods=['POST'])
def run_reconciliation_matching():
    """Automatic bank matching disabled — not required for launch reconciliation workflow."""
    return redirect(url_for('reconciliation'))


@app.route('/reconciliation/apply-fix', methods=['POST'])
@require_admin
def apply_reconciliation_fix():
    """Bank matching fix actions disabled — not required for launch reconciliation workflow."""
    return redirect(url_for('reconciliation'))


@app.route('/reconciliation/clear-bank-session', methods=['POST'])
def clear_bank_session():
    """Bank session management disabled — not required for launch reconciliation workflow."""
    return redirect(url_for('reconciliation'))


@app.route('/audit-log')
@require_admin
def audit_log():
    """Audit log page (admin only)."""
    username_filter = request.args.get('username', '')
    module_filter = request.args.get('module', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    entries = db.get_audit_log_entries(
        username=username_filter or None,
        module=module_filter or None,
        date_from=date_from or None,
        date_to=date_to or None
    )
    usernames = db.get_audit_usernames()
    modules = db.get_audit_modules()
    return render_template('audit_log.html',
                         entries=entries,
                         usernames=usernames,
                         modules=modules,
                         username_filter=username_filter,
                         module_filter=module_filter,
                         date_from=date_from,
                         date_to=date_to)


def _get_task_status_safe() -> dict:
    """Return task_scheduler status dict, failing safely to 'unknown' state."""
    try:
        from task_scheduler import get_task_status
        return get_task_status()
    except Exception:
        return {
            'exists': None,  # None = could not determine
            'state': 'Unknown',
            'last_run': None, 'next_run': None,
            'last_result': None, 'run_as': None,
            'system_level': False,
        }


@app.route('/system-backups')
@require_admin
def system_backups():
    """System backups page (admin only)."""
    from backup_service import (
        get_last_backup_time,
        get_onedrive_backup_dir,
        get_active_cloud_backup_dir,
        get_custom_backup_dir_raw,
        LOCAL_BACKUP_DIR,
        USB_BACKUP_INTERVAL_DAYS,
    )
    from datetime import datetime, timedelta

    last_backup = get_last_backup_time()
    last_usb_str = db.get_config('last_usb_backup_date', '')
    last_usb = None
    if last_usb_str:
        try:
            last_usb = datetime.strptime(last_usb_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    usb_status = 'Completed'
    usb_due = False
    if last_usb:
        days_since = (datetime.now().date() - last_usb).days
        if days_since >= USB_BACKUP_INTERVAL_DAYS:
            usb_status = 'Due'
            usb_due = True
    else:
        usb_status = 'Due'
        usb_due = True

    # Determine active cloud backup destination and its type
    od_path = get_onedrive_backup_dir()
    custom_raw = get_custom_backup_dir_raw(db.db_path)
    active_cloud = get_active_cloud_backup_dir(db.db_path)

    if custom_raw and os.path.isdir(custom_raw):
        cloud_type = 'custom'
        cloud_dir = custom_raw
    elif active_cloud and active_cloud == od_path:
        cloud_type = 'onedrive'
        cloud_dir = od_path
    else:
        cloud_type = 'none'
        cloud_dir = None

    return render_template('system_backups.html',
                         last_backup=last_backup,
                         local_backup_dir=LOCAL_BACKUP_DIR,
                         onedrive_backup_dir=od_path or 'OneDrive not detected',
                         cloud_dir=cloud_dir,
                         cloud_type=cloud_type,
                         custom_dir_set=bool(custom_raw),
                         last_usb_date=last_usb_str or 'Never',
                         usb_status=usb_status,
                         usb_due=usb_due,
                         task_status=_get_task_status_safe())


@app.route('/system-backups/create-now', methods=['POST'])
@require_admin
def backup_create_now():
    """Create backup immediately (admin only)."""
    nexal_data = os.environ.get('NEXAL_DATA_DIR', '').strip()
    if nexal_data:
        try:
            from nexal_platform.backup import BackupService

            result = BackupService().run_backup(schedule='daily')
            if result.success:
                db.set_config('last_backup_failure', '')
                flash(f'Multi-tenant backup OK: {result.manifest_path}', 'success')
            else:
                db.set_config(
                    'last_backup_failure',
                    datetime.now().strftime('%Y-%m-%d %H:%M') + ': ' + (result.error or 'failed'),
                )
                flash(f'Backup failed: {result.error}', 'error')
        except Exception as exc:
            db.set_config(
                'last_backup_failure',
                datetime.now().strftime('%Y-%m-%d %H:%M') + ': ' + str(exc),
            )
            flash(f'Backup failed: {exc}', 'error')
        return redirect(url_for('system_backups'))

    from backup_service import create_backup

    def audit(action, details):
        log_audit('Backup System', action, details=details)

    success, msg, _ = create_backup(db_path=db.db_path, audit_callback=audit)
    if success:
        db.set_config('last_backup_failure', '')
        flash(msg, 'success')
    else:
        db.set_config('last_backup_failure', datetime.now().strftime('%Y-%m-%d %H:%M') + ': ' + str(msg))
        flash(f'Backup failed: {msg}', 'error')
    return redirect(url_for('system_backups'))


@app.route('/system-backups/validate-path', methods=['POST'])
@require_admin
def validate_backup_path():
    """Validate a candidate backup path (exists + writable). Returns JSON. No state changes."""
    path = (request.form.get('path') or '').strip()
    if not path:
        return {'ok': False, 'exists': False, 'writable': False, 'error': 'No path provided.'}
    exists = os.path.isdir(path)
    if not exists:
        return {'ok': False, 'exists': False, 'writable': False, 'error': f'Folder not found: {path}'}
    test_file = os.path.join(path, '.solicitor_write_test')
    try:
        with open(test_file, 'w') as fh:
            fh.write('write_test')
        os.remove(test_file)
        return {'ok': True, 'exists': True, 'writable': True}
    except Exception as e:
        return {'ok': False, 'exists': True, 'writable': False, 'error': f'Folder is not writable: {e}'}


@app.route('/system-backups/repair-scheduler', methods=['POST'])
@require_admin
def repair_backup_scheduler():
    """Re-register the Windows Task Scheduler task (admin only)."""
    try:
        from task_scheduler import repair_task
        success, msg, system_level = repair_task()
        if success:
            log_audit('Backup System', 'BACKUP_TASK_REPAIRED', details=msg)
            flash(msg, 'success')
        else:
            flash(f'Scheduler repair failed: {msg}', 'error')
    except Exception as exc:
        flash(f'Scheduler repair error: {exc}', 'error')
    return redirect(url_for('system_backups'))


@app.route('/system-backups/pick-folder', methods=['POST'])
@require_admin
def pick_backup_folder():
    """Open native Windows folder picker; return selected path as JSON."""
    import subprocess
    try:
        ps_script = (
            "Add-Type -AssemblyName System.Windows.Forms; "
            "$d = New-Object System.Windows.Forms.FolderBrowserDialog; "
            "$d.Description = 'Select Backup Destination Folder'; "
            "$d.ShowNewFolderButton = $true; "
            "$owner = New-Object System.Windows.Forms.Form; "
            "$owner.TopMost = $true; "
            "$owner.WindowState = 'Minimized'; "
            "$owner.Show(); "
            "$r = $d.ShowDialog($owner); "
            "$owner.Close(); "
            "if ($r -eq 'OK') { Write-Output $d.SelectedPath }"
        )
        flags = 0x08000000 if os.name == 'nt' else 0  # CREATE_NO_WINDOW
        result = subprocess.run(
            ['powershell', '-WindowStyle', 'Hidden', '-NonInteractive', '-Command', ps_script],
            capture_output=True, text=True, timeout=120,
            creationflags=flags
        )
        path = (result.stdout or '').strip()
        if path and os.path.isdir(path):
            return {'success': True, 'path': path}
        return {'success': False, 'error': 'No folder selected.'}
    except subprocess.TimeoutExpired:
        return {'success': False, 'error': 'Folder picker timed out.'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


@app.route('/system-backups/set-backup-location', methods=['POST'])
@require_admin
def set_backup_location():
    """Save or clear the custom cloud backup destination (admin only)."""
    from backup_service import set_custom_backup_dir, clear_custom_backup_dir
    action = request.form.get('action', 'set')

    if action == 'clear':
        clear_custom_backup_dir(db.db_path)
        log_audit('Backup System', 'BACKUP_LOCATION_CLEARED',
                  details='Custom backup location removed — reverted to OneDrive auto-detection')
        flash('Backup location reset to automatic OneDrive detection.', 'success')
        return redirect(url_for('system_backups'))

    path = (request.form.get('backup_path') or '').strip()
    if not path:
        flash('No path provided.', 'error')
        return redirect(url_for('system_backups'))

    # Validate folder exists
    if not os.path.isdir(path):
        flash(f'Folder not found: {path}', 'error')
        return redirect(url_for('system_backups'))

    # Test write permission
    test_file = os.path.join(path, '.solicitor_write_test')
    try:
        with open(test_file, 'w') as fh:
            fh.write('write_test')
        os.remove(test_file)
    except Exception as e:
        flash(f'Folder is not writable: {e}', 'error')
        return redirect(url_for('system_backups'))

    set_custom_backup_dir(db.db_path, path)
    log_audit('Backup System', 'BACKUP_LOCATION_SET',
              details=f'Cloud backup destination set to: {path}')
    flash(f'Backup destination saved: {path}', 'success')
    return redirect(url_for('system_backups'))


@app.route('/system-backups/restore', methods=['GET', 'POST'])
@require_admin
def restore_backup_route():
    """
    Restore database from backup file (admin only).
    Supports both encrypted (.zip.enc) and legacy (.zip) backups.
    """
    from backup_service import restore_backup as do_restore, ENCRYPTED_EXTENSION
    
    if request.method == 'GET':
        return redirect(url_for('system_backups'))
    if 'backup_file' not in request.files or not request.files['backup_file'].filename:
        flash('No backup file selected.', 'error')
        return redirect(url_for('system_backups'))
    
    f = request.files['backup_file']
    fname_lower = f.filename.lower()
    
    # Accept both .zip and .zip.enc files
    if not (fname_lower.endswith('.zip') or fname_lower.endswith(ENCRYPTED_EXTENSION)):
        flash('Invalid file. Select a backup file (.zip or .zip.enc).', 'error')
        return redirect(url_for('system_backups'))
    
    try:
        import tempfile
        import shutil
        
        # Save uploaded file to temp location
        temp_dir = tempfile.mkdtemp()
        try:
            temp_backup = os.path.join(temp_dir, f.filename)
            f.save(temp_backup)
            
            # Use backup_service restore function (handles encryption)
            success, msg = do_restore(temp_backup, db.db_path, 
                                      audit_callback=lambda a, d: log_audit('Backup System', a, details=d))
            
            if success:
                flash(msg, 'success')
            else:
                flash(msg, 'error')
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)
            
    except Exception as e:
        flash(f'Restore failed: {e}', 'error')
        log_audit('Backup System', 'RESTORE_FAILED', details=str(e))
    
    return redirect(url_for('system_backups'))


@app.route('/system-backups/export-compliance-pack')
@require_admin
def export_compliance_pack():
    """Generate ZIP with audit log, reconciliation, cashbook, ledger, backup info (admin only)."""
    import zipfile
    from io import BytesIO
    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        entries = db.get_audit_log_entries(limit=5000)
        lines = ['timestamp,username,role,action,module,record_id,details']
        for e in entries:
            lines.append(','.join([
                str(e.get('timestamp', '')),
                str(e.get('username', '')).replace(',', ';'),
                str(e.get('role', '')),
                str(e.get('action', '')).replace(',', ';'),
                str(e.get('module', '')),
                str(e.get('record_id', '')),
                '"' + str(e.get('details', '')).replace('"', '""') + '"'
            ]))
        zf.writestr('audit_log.csv', '\n'.join(lines))
        recs = db.get_most_recent_reconciliation_date()
        zf.writestr('reconciliation_summary.txt', f"Last reconciliation: {recs or 'None'}\n")
        trans = db.get_all_cashbook_transactions()
        clines = ['Transaction ID,Date,Type,Reference,Source,Status,Amount']
        for t in trans:
            clines.append(','.join([str(t.get('transaction_id', '')), str(t.get('transaction_date', '')), str(t.get('transaction_type', '')), str(t.get('reference', '')).replace(',', ';'), str(t.get('source', '')), str(t.get('status', '')), str(t.get('amount', ''))]))
        zf.writestr('cashbook_report.csv', '\n'.join(clines))
        clients = db.get_all_clients()
        blines = ['Client Code,Client Name,Balance']
        for c in clients:
            bal = db.get_client_balance(c['id'])
            blines.append(f"{c.get('client_code','')},{c.get('client_name','').replace(',',';')},{bal}")
        zf.writestr('client_ledger_balances.csv', '\n'.join(blines))
        from backup_service import get_last_backup_time, LOCAL_BACKUP_DIR
        last = get_last_backup_time()
        zf.writestr('backup_verification.txt', f"Last backup: {last}\nLocal folder: {LOCAL_BACKUP_DIR}\n")
    buf.seek(0)
    fn = f"compliance_pack_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
    log_audit('Compliance', 'Compliance pack exported', details=fn)
    if request.args.get('save_local'):
        return _save_local_response(buf.read(), fn)
    return send_file(buf, mimetype='application/zip', as_attachment=True, download_name=fn)


@app.route('/admin/reset-january', methods=['GET', 'POST'])
@require_admin
def admin_reset_january():
    """
    Admin-only: soft-reset all January 2026 client-money activity and validate against December closing.
    Requires ALLOW_JANUARY_RESET=true and typed confirmation RESET JANUARY.
    """
    from decimal import Decimal as D

    year, month = 2026, 1
    expected_client_money = D('168736.32')
    require_locked = [(11, 2025), (12, 2025)]
    if not (is_january_2026('2026-01-01') and is_january_2026('2026-01-31')):
        flash('Reset window configuration is invalid.', 'error')
        return redirect(url_for('reconciliation'))

    if request.method == 'POST':
        phrase = (request.form.get('confirm_phrase') or '').strip()
        if phrase != 'RESET JANUARY':
            flash('Confirmation phrase must be exactly RESET JANUARY.', 'error')
            return redirect(url_for('admin_reset_january'))
        try:
            counts = db.reset_unlocked_calendar_month_client_money(
                year,
                month,
                expected_client_money,
                True,
                current_username(),
                require_months_locked=require_locked,
            )
            log_audit(
                'System',
                'JANUARY_MONTH_SOFT_RESET',
                details=(
                    f"user={current_username()} counts={counts} "
                    f"verified_ledger={counts.get('verified_ledger_net')} "
                    f"verified_cashbook={counts.get('verified_cashbook_net')}"
                ),
            )
            flash(
                'January reset successfully (soft reset). Opening balance restored to £168,736.32. '
                'January rows are retained for audit but excluded from operational views/reports.',
                'success',
            )
        except RuntimeError as e:
            flash(str(e), 'error')
        except ValueError as e:
            flash(str(e), 'error')
        except Exception as e:
            logger.exception('January reset failed')
            flash(f'Reset failed: {e}', 'error')
        return redirect(url_for('reconciliation'))

    return render_template(
        'admin_reset_january.html',
        expected_balance=expected_client_money,
        jan_locked=db.is_month_locked(month, year),
        lock_nov=db.is_month_locked(11, 2025),
        lock_dec=db.is_month_locked(12, 2025),
    )


@app.route('/admin/audit-ledger/<int:client_id>')
@require_admin
def admin_audit_ledger(client_id):
    """
    Read-only forensic audit: compares per-step running total (Decimal, 2dp after each row)
    to the database-authoritative cumulative balance through the same row (date, id order).
    Does not modify data.
    """
    client = db.get_client(client_id)
    if not client:
        flash('Client not found.', 'error')
        return redirect(url_for('client_ledger'))
    report = db.audit_client_ledger_running_balance(client_id)
    logger.info(
        'Ledger forensic audit client_id=%s rows=%s discrepancy=%s final_ok=%s',
        client_id,
        report.get('row_count'),
        report.get('first_discrepancy') is not None,
        report.get('final_ok'),
    )
    return render_template('audit_ledger_client.html', client=client, report=report)


@app.route('/admin/reset-system', methods=['GET', 'POST'])
@require_admin
def admin_reset_system():
    """
    Admin-only: wipe all financial data (and optionally all clients). Preserves user accounts.
    Requires ALLOW_FULL_SYSTEM_RESET=true; blocked if SOLICITOR_NO_RESET is set.
    """
    allow = os.environ.get('ALLOW_FULL_SYSTEM_RESET', '').strip().lower() == 'true'
    blocked_no_reset = bool(os.environ.get('SOLICITOR_NO_RESET', ''))
    env_default_delete_clients = os.environ.get('RESET_DELETE_CLIENTS', 'true').strip().lower() == 'true'

    if request.method == 'POST':
        if not allow or blocked_no_reset:
            flash('Full system reset is not allowed (check ALLOW_FULL_SYSTEM_RESET and SOLICITOR_NO_RESET).', 'error')
            return redirect(url_for('admin_reset_system'))
        phrase = (request.form.get('confirm_phrase') or '').strip()
        if phrase != 'RESET SYSTEM':
            flash('Confirmation phrase must be exactly RESET SYSTEM.', 'error')
            return redirect(url_for('admin_reset_system'))
        delete_clients = request.form.get('delete_clients') == 'on'
        try:
            counts = db.full_system_reset(True, delete_clients, current_username())
            log_audit(
                'System',
                'SYSTEM_RESET',
                details=(
                    f"user={current_username()} delete_clients={delete_clients} counts={counts}"
                ),
            )
            flash('System reset successfully. All data cleared.', 'success')
        except RuntimeError as e:
            flash(str(e), 'error')
        except ValueError as e:
            flash(str(e), 'error')
        except Exception as e:
            logger.exception('Full system reset failed')
            flash(f'Reset failed: {e}', 'error')
        return redirect(url_for('index'))

    return render_template(
        'admin_reset_system.html',
        allow_full_reset=allow,
        blocked_by_no_reset=blocked_no_reset,
        default_delete_clients=env_default_delete_clients,
    )


@app.route('/admin/reset-all', methods=['POST'])
@require_admin
def admin_reset_all():
    """
    Hard reset to fresh-install state.
    Permanently wipes all operational data and resets sequences.
    """
    allowed = os.environ.get('ALLOW_FULL_SYSTEM_RESET', '').strip().lower() == 'true'
    blocked = bool(os.environ.get('SOLICITOR_NO_RESET', ''))
    if not allowed or blocked:
        flash('Hard reset is blocked (check ALLOW_FULL_SYSTEM_RESET and SOLICITOR_NO_RESET).', 'error')
        return redirect(url_for('system_health'))
    try:
        db.reset_database(confirm=True)
        session.clear()
        return 'System fully reset to brand new state'
    except RuntimeError as e:
        flash(str(e), 'error')
    except Exception as e:
        logger.exception('Hard reset failed')
        flash(f'Hard reset failed: {e}', 'error')
    return redirect(url_for('system_health'))


@app.route('/system-health')
@require_admin
def system_health():
    """System health status (admin only)."""
    from backup_service import get_last_backup_time, get_onedrive_backup_dir, LOCAL_BACKUP_DIR
    last_backup = get_last_backup_time()
    last_usb = db.get_config('last_usb_backup_date', '')
    last_rec = db.get_most_recent_reconciliation_date()
    integrity_ok, integrity_msg = db.verify_database_integrity()
    consistency_err = db.verify_ledger_consistency()
    return render_template('system_health.html',
                         last_backup=last_backup,
                         last_usb_date=last_usb or 'Never',
                         last_reconciliation=last_rec or 'Never',
                         integrity_ok=integrity_ok,
                         integrity_msg=integrity_msg,
                         consistency_ok=consistency_err is None,
                         consistency_error=consistency_err,
                         local_backup_dir=LOCAL_BACKUP_DIR,
                         onedrive_dir=get_onedrive_backup_dir() or 'Not detected')


@app.route('/admin/security')
@require_admin
def admin_security():
    """Account information — authentication is managed in the Portal."""
    return render_template('security.html')


@app.route('/admin/security/generate-recovery-key', methods=['POST'])
@require_admin
def admin_generate_recovery_key():
    return redirect(url_for('admin_security'))


@app.route('/admin/security/recovery-key-ack', methods=['POST'])
@require_admin
def admin_recovery_key_ack():
    return redirect(url_for('admin_security'))


@app.route('/admin/recovery', methods=['GET', 'POST'])
def admin_recovery():
    return portal_login_redirect(reason='legacy_route')


@app.route('/admin/recovery/reset', methods=['GET', 'POST'])
def admin_recovery_reset():
    return portal_login_redirect(reason='legacy_route')


@app.route('/user-management', methods=['GET'])
@require_admin
def user_management():
    """Admin-only user management page."""
    from lib.firm_package import package_usage_summary

    users = db.get_billable_users_for_management()
    package_usage = package_usage_summary(db, session)
    return render_template(
        'user_management.html',
        users=users,
        current_user_id=session.get('user_id'),
        package_usage=package_usage,
    )


@app.route('/user-management/add-user', methods=['POST'])
@require_admin
def user_management_add_user():
    flash('User invitations and account management are handled through the Nexal Legal Portal.', 'info')
    return redirect(url_for('user_management'))


@app.route('/admin/reset-password/<int:user_id>', methods=['POST'])
@require_admin
def admin_reset_password(user_id):
    return redirect(url_for('user_management'))


@app.route('/user-management/reset-link/<token>', methods=['GET'])
@require_admin
def user_management_reset_link_page(token):
    return portal_login_redirect(reason='legacy_route')


@app.route('/user-management/change-role/<int:user_id>', methods=['POST'])
@require_admin
def user_management_change_role(user_id):
    """Change user role to admin."""
    user = db.get_user_by_id(user_id)
    if not user:
        flash('User not found.', 'error')
        return redirect(url_for('user_management'))
    db.update_user_role(user_id, 'admin')
    log_audit('User Management', 'Role changed to Admin', record_id=str(user_id), details=user['username'])
    flash(f'User {user["username"]} is now an Admin.', 'success')
    return redirect(url_for('user_management'))


@app.route('/user-management/deactivate/<int:user_id>', methods=['POST'])
@require_admin
def user_management_deactivate(user_id):
    """Deactivate user. Admin cannot deactivate self."""
    if user_id == session.get('user_id'):
        flash('You cannot deactivate your own account.', 'error')
        return redirect(url_for('user_management'))
    user = db.get_user_by_id(user_id)
    if not user:
        flash('User not found.', 'error')
        return redirect(url_for('user_management'))
    db.set_user_active(user_id, False)
    log_audit('User Management', 'User deactivated', record_id=str(user_id), details=user['username'])
    flash(f'User {user["username"]} has been deactivated.', 'success')
    return redirect(url_for('user_management'))


@app.route('/user-management/regenerate-recovery-key', methods=['POST'])
@require_admin
def user_management_regenerate_recovery_key():
    return redirect(url_for('user_management'))


@app.route('/admin-reset-password/<token>', methods=['GET', 'POST'])
def admin_reset_password_page(token):
    return portal_login_redirect(reason='legacy_route')


@app.route('/reset-password/<token>', methods=['GET'])
def reset_password(token):
    return portal_login_redirect(reason='legacy_route')


@app.route('/force-password-change', methods=['GET', 'POST'])
def force_password_change():
    session.clear()
    return portal_login_redirect(reason='legacy_route')


@app.route('/system-backups/usb', methods=['POST'])
@require_admin
def backup_to_usb():
    """Copy latest backup to USB (admin only)."""
    from backup_service import copy_to_usb
    from datetime import datetime

    def audit(action, details):
        log_audit('Backup System', action, details=details)

    success, msg = copy_to_usb(audit_callback=audit)
    if success:
        db.set_config('last_usb_backup_date', datetime.now().strftime('%Y-%m-%d'))
        flash(msg, 'success')
    else:
        flash(msg, 'error')
    return redirect(url_for('system_backups'))


def _client_cashbook_total_to_date(end_date: str = None) -> Decimal:
    """Sum cleared client-linked cashbook (client money only) to end_date inclusive (or all if None)."""
    client_cashbook = db.get_all_cashbook_transactions(end_date=end_date, client_only=True)
    total = Decimal('0')
    for trans in client_cashbook:
        is_cleared = trans['status'] == 'Cleared'
        is_not_reversed = (trans.get('reversal_status') or 'ACTIVE') != 'REVERSED'
        depth_ok = (trans.get('reversal_depth') or 0) % 2 == 0
        if is_cleared and is_not_reversed and depth_ok:
            if trans['transaction_type'] == 'Receipt':
                total += Decimal(str(trans['amount']))
            else:
                total -= Decimal(str(trans['amount']))
    return total


@app.route('/reconciliation/new', methods=['GET', 'POST'])
def new_reconciliation():
    """
    Create new CLIENT MONEY reconciliation.
    
    SRA Accounts Rules require reconciliation of CLIENT money only.
    Office account balances are excluded - they are reported separately.
    
    The reconciliation verifies:
    - Client Ledger Total = sum of all client matter balances
    - Client Cashbook Total = sum of all client money transactions
    - Client Bank Balance = actual bank balance for client account
    - All three should match when properly reconciled (within £0.01 tolerance)
    """
    from reconciliation_utils import compute_reconciliation_state

    if request.method == 'POST':
        try:
            rec_date = parse_transaction_date_strict(
                request.form.get('reconciliation_date'), 'reconciliation_date'
            )
            log_transaction_date_saved(rec_date, 'reconciliation')
            mw = month_outside_current_warning(rec_date)
            if mw:
                flash(mw, 'warning')

            ledger_total = db.get_total_ledger_balance(as_of_date=rec_date)
            cashbook_total = _client_cashbook_total_to_date(rec_date)
            bank_balance = Decimal(request.form['bank_balance'])
            notes = request.form.get('notes') or None

            state = compute_reconciliation_state(ledger_total, cashbook_total, bank_balance)
            if not state['can_complete']:
                flash(
                    'Cannot complete reconciliation: Ledger, Cashbook, and bank balance must agree '
                    'within £0.01 before recording.',
                    'error',
                )
                return render_template(
                    'new_reconciliation.html',
                    today=rec_date,
                    recon=state,
                    notes_prefill=notes or '',
                )
            
            rec_id = db.create_reconciliation(
                reconciliation_date=rec_date,
                ledger_total=ledger_total,
                cashbook_total=cashbook_total,
                bank_balance=bank_balance,
                notes=notes
            )
            log_audit('Reconciliation', 'CLIENT_RECONCILIATION_COMPLETED', record_id=str(rec_id),
                     details=f"{rec_date} client_ledger={ledger_total} client_cashbook={cashbook_total} client_bank={bank_balance}")
            flash(f'Client Money Reconciliation created successfully for {rec_date}', 'success')
            return redirect(url_for('reconciliation'))
        except ValueError as e:
            flash(str(e), 'error')
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
    
    today = request.args.get('date') or datetime.now().strftime('%Y-%m-%d')
    try:
        parse_transaction_date_strict(today, 'date')
    except ValueError:
        today = datetime.now().strftime('%Y-%m-%d')
    ledger_total = db.get_total_ledger_balance(as_of_date=today)
    bank_balance = db.get_bank_balance(as_of_date=today, client_only=True)
    cashbook_total = _client_cashbook_total_to_date(today)
    recon = compute_reconciliation_state(ledger_total, cashbook_total, bank_balance)

    return render_template(
        'new_reconciliation.html',
        today=today,
        recon=recon,
        notes_prefill='',
    )


@app.route('/reports')
def reports():
    """Reports page"""
    clients = db.get_all_clients()
    users = db.get_billable_active_users()
    return render_template('reports.html', clients=clients, users=users)


@app.route('/api/client-balance/<int:client_id>')
def api_client_balance(client_id):
    """API endpoint for client balance"""
    balance = db.get_client_balance(client_id)
    return jsonify({'balance': str(balance)})


def build_pdf_report(title, subtitle, headers, rows, col_widths, currency_cols=None):
    """
    Reusable PDF report builder for all exports.
    - Landscape A4
    - Consistent header/table/footer styling
    - Wrapped cells via Paragraph
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=26,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    cell_style = styles['Normal'].clone('PdfCell')
    cell_style.fontName = 'Helvetica'
    cell_style.fontSize = 8
    cell_style.leading = 10

    from lib.branding import LEGAL_COMPANY_NAME, PRODUCT_NAME

    elements = []
    elements.append(Paragraph(PRODUCT_NAME, styles['Heading2']))
    elements.append(Paragraph(LEGAL_COMPANY_NAME, styles['Normal']))
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    if subtitle:
        elements.append(Paragraph(subtitle, styles['Normal']))
    elements.append(Spacer(1, 10))

    if not rows:
        elements.append(Paragraph("No data in selected range.", styles['Normal']))
    else:
        table_data = []
        table_data.append([Paragraph(str(h), cell_style) for h in headers])
        for row in rows:
            table_data.append([Paragraph(str(v if v is not None else ''), cell_style) for v in row])

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        ts = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]
        for c in (currency_cols or []):
            ts.append(('ALIGN', (c, 1), (c, -1), 'RIGHT'))
        table.setStyle(TableStyle(ts))
        elements.append(table)

    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        f"Generated by Nexal Legal — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        styles['Normal'],
    ))
    doc.build(elements)
    buffer.seek(0)
    return buffer


def _build_ledger_pdf(transactions, date_from, date_to, client_filter):
    """Generate PDF for client ledger report."""
    date_range = f"{date_from or 'Start'} to {date_to or 'End'}"
    rows = []
    receipts_total = Decimal('0')
    payments_total = Decimal('0')
    for t in transactions:
        amt = Decimal(str(t['amount']))
        if t['transaction_type'] == 'Receipt':
            receipts_total += amt
        else:
            payments_total += amt
        rows.append([
            str(t.get('transaction_id', '-')),
            str(t['transaction_date']),
            f"{t.get('client_code', '')} - {t.get('client_name', '')}" if t.get('client_code') else '-',
            str(t['transaction_type']),
            str(t.get('reference') or '-'),
            str(t['source']),
            f"£{amt:,.2f}" if t['transaction_type'] == 'Receipt' else f"-£{amt:,.2f}",
            str(t.get('created_by') or 'System'),
        ])
    rows.append(['', '', '', '', '', 'Opening:', '£0.00', ''])
    rows.append(['', '', '', '', '', 'Receipts:', f"£{receipts_total:,.2f}", ''])
    rows.append(['', '', '', '', '', 'Payments:', f"£{payments_total:,.2f}", ''])
    rows.append(['', '', '', '', '', 'Closing:', f"£{(receipts_total - payments_total):,.2f}", ''])
    return build_pdf_report(
        "Client Ledger Report",
        f"Date Range: {date_range} | Client: {client_filter}",
        ['TXN ID', 'Date', 'Client', 'Type', 'Reference', 'Source', 'Amount', 'Created By'],
        rows,
        [90, 70, 140, 70, 160, 90, 80, 60],
        currency_cols=[6],
    )


def _build_cashbook_pdf(transactions, date_from, date_to):
    """Generate PDF for cashbook report."""
    date_range = f"{date_from or 'Start'} to {date_to or 'End'}"
    rows = []
    receipts_total = Decimal('0')
    payments_total = Decimal('0')
    running = Decimal('0')
    chronological = sorted(transactions, key=lambda x: (x['transaction_date'], x['id']))
    for t in chronological:
        if t['status'] == 'Declined':
            continue
        amt = Decimal(str(t['amount']))
        if t['transaction_type'] == 'Receipt':
            receipts_total += amt
            running += amt
        else:
            payments_total += amt
            running -= amt
        client_disp = (t.get('client_code') or '') + (' - ' + (t.get('client_name') or '') if t.get('client_code') else (t.get('client_name') or 'Standalone'))
        rows.append([
            str(t.get('transaction_id', '-')),
            str(t['transaction_date']),
            str(t['transaction_type']),
            str(t.get('reference') or '-'),
            str(t['source']),
            str(t['status']),
            client_disp,
            f"£{amt:,.2f}" if t['transaction_type'] == 'Receipt' else f"-£{amt:,.2f}",
            f"£{running:,.2f}",
            str(t.get('created_by', 'System')),
        ])
    rows.append(['', '', '', '', '', '', 'Opening:', '£0.00', '£0.00', ''])
    rows.append(['', '', '', '', '', '', 'Receipts:', f"£{receipts_total:,.2f}", '', ''])
    rows.append(['', '', '', '', '', '', 'Payments:', f"£{payments_total:,.2f}", '', ''])
    rows.append(['', '', '', '', '', '', 'Closing:', f"£{(receipts_total - payments_total):,.2f}", f"£{(receipts_total - payments_total):,.2f}", ''])
    return build_pdf_report(
        "Cashbook Report",
        f"Date Range: {date_range}",
        ['TXN ID', 'Date', 'Type', 'Reference', 'Source', 'Status', 'Client', 'Amount', 'Running Balance', 'Created By'],
        rows,
        [75, 60, 55, 130, 70, 60, 120, 70, 80, 55],
        currency_cols=[7, 8],
    )


@app.route('/reports/export/ledger-pdf')
def export_ledger_pdf():
    """Export client ledger report as PDF."""
    if not REPORTLAB_AVAILABLE:
        return jsonify({'success': False, 'error': 'PDF library (reportlab) is not installed.'}), 500
    try:
        client_id = request.args.get('client_id', type=int)
        date_from = request.args.get('date_from') or None
        date_to = request.args.get('date_to') or None
        client_filter = "All Clients" if not client_id else ""
        if client_id:
            client = db.get_client(client_id)
            if client:
                client_filter = f"{client['client_code']} - {client['client_name']}"
        created_by = request.args.get('created_by') or None
        transactions = db.get_ledger_transactions_for_report(client_id, date_from, date_to, created_by)
        buffer = _build_ledger_pdf(transactions, date_from, date_to, client_filter)
        fn = f"ledger_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=fn,
            max_age=0,
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reports/export/ledger-csv')
def export_ledger_csv():
    """Export client ledger report as CSV."""
    client_id = request.args.get('client_id', type=int)
    date_from = request.args.get('date_from') or None
    date_to = request.args.get('date_to') or None
    created_by = request.args.get('created_by') or None
    transactions = db.get_ledger_transactions_for_report(client_id, date_from, date_to, created_by)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['Transaction ID', 'Date', 'Client Code', 'Client Name', 'Type', 'Reference', 'Source', 'Amount', 'Created By'])
    for t in transactions:
        writer.writerow([
            t.get('transaction_id', ''),
            t['transaction_date'],
            t.get('client_code', ''),
            t.get('client_name', ''),
            t['transaction_type'],
            t['reference'],
            t['source'],
            t['amount'],
            t.get('created_by', 'System'),
        ])
    buffer.seek(0)
    fn = f"ledger_report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    data = buffer.getvalue().encode('utf-8-sig')
    if request.args.get('save_local'):
        return _save_local_response(data, fn)
    return send_file(io.BytesIO(data), mimetype='text/csv', as_attachment=True, download_name=fn)


@app.route('/reports/export/cashbook-pdf')
def export_cashbook_pdf():
    """Export cashbook report as PDF."""
    if not REPORTLAB_AVAILABLE:
        return jsonify({'success': False, 'error': 'PDF library (reportlab) is not installed.'}), 500
    try:
        date_from = request.args.get('date_from') or None
        date_to = request.args.get('date_to') or None
        created_by = request.args.get('created_by') or None
        transactions = db.get_all_cashbook_transactions(start_date=date_from, end_date=date_to, created_by=created_by)
        buffer = _build_cashbook_pdf(transactions, date_from, date_to)
        fn = f"cashbook_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=fn,
            max_age=0,
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reports/export/cashbook-csv')
def export_cashbook_csv():
    """Export cashbook report as CSV."""
    date_from = request.args.get('date_from') or None
    date_to = request.args.get('date_to') or None
    created_by = request.args.get('created_by') or None
    transactions = db.get_all_cashbook_transactions(start_date=date_from, end_date=date_to, created_by=created_by)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['Transaction ID', 'Date', 'Type', 'Reference', 'Source', 'Status', 'Client', 'Amount', 'Created By'])
    for t in transactions:
        client = f"{t.get('client_code', '')} - {t.get('client_name', '')}".strip(' -') or 'Standalone'
        writer.writerow([
            t.get('transaction_id', ''),
            t['transaction_date'],
            t['transaction_type'],
            t['reference'],
            t['source'],
            t['status'],
            client,
            t['amount'],
            t.get('created_by', 'System'),
        ])
    buffer.seek(0)
    fn = f"cashbook_report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    data = buffer.getvalue().encode('utf-8-sig')
    if request.args.get('save_local'):
        return _save_local_response(data, fn)
    return send_file(io.BytesIO(data), mimetype='text/csv', as_attachment=True, download_name=fn)


def _build_office_income_pdf(transactions, date_from, date_to, total):
    """Generate PDF for office income report."""
    date_range = f"{date_from or 'Start'} to {date_to or 'End'}"
    rows = []
    running = Decimal('0')
    chronological = sorted(transactions, key=lambda x: (x['transaction_date'], str(x.get('id', ''))))
    for t in chronological:
        amt = Decimal(str(t['amount']))
        running += amt
        if t.get('transaction_type') == 'Fee Transfer':
            client = f"{t.get('client_code', '')} - {t.get('client_name', '')}" if t.get('client_code') else '-'
            typ = 'Fee Transfer'
        else:
            client = '-'
            typ = t.get('transaction_type') or 'Receipt'
        rows.append([str(t['transaction_date']), typ, str(t.get('reference') or '-'), client, f"£{amt:,.2f}", f"£{running:,.2f}", str(t.get('created_by', 'System'))])
    rows.append(['', '', '', 'Opening:', '£0.00', '£0.00', ''])
    rows.append(['', '', '', 'Closing:', f"£{total:,.2f}", f"£{total:,.2f}", ''])
    return build_pdf_report(
        "Office Income Summary",
        f"Date Range: {date_range}",
        ['Date', 'Type', 'Reference', 'Client', 'Amount', 'Running Balance', 'Created By'],
        rows,
        [70, 80, 180, 200, 80, 90, 60],
        currency_cols=[4, 5],
    )


def _build_office_expenses_pdf(transactions, date_from, date_to, total):
    """Generate PDF for office expenses report."""
    date_range = f"{date_from or 'Start'} to {date_to or 'End'}"
    rows = []
    running = Decimal('0')
    chronological = sorted(transactions, key=lambda x: (x['transaction_date'], str(x.get('id', ''))))
    for t in chronological:
        amt = Decimal(str(t['amount']))
        running -= amt
        rows.append([
            str(t['transaction_date']),
            str(t.get('reference') or '-'),
            str(t.get('source') or '-'),
            str(t.get('description') or '-'),
            f"£{amt:,.2f}",
            f"£{running:,.2f}",
            str(t.get('created_by', 'System')),
        ])
    rows.append(['', '', '', 'Opening:', '£0.00', '£0.00', ''])
    rows.append(['', '', '', 'Closing:', f"-£{total:,.2f}", f"-£{total:,.2f}", ''])
    return build_pdf_report(
        "Office Expenses Summary",
        f"Date Range: {date_range}",
        ['Date', 'Reference', 'Source', 'Description', 'Amount', 'Running Balance', 'Created By'],
        rows,
        [70, 130, 80, 250, 80, 90, 60],
        currency_cols=[4, 5],
    )


def _build_office_profit_pdf(income_total, expenses_total, date_from, date_to):
    """Generate PDF for net profit report."""
    date_range = f"{date_from or 'Start'} to {date_to or 'End'}"
    net = income_total - expenses_total
    rows = [
        ['Opening Balance', '£0.00'],
        ['Office Income', f"£{income_total:,.2f}"],
        ['Office Expenses', f"£{expenses_total:,.2f}"],
        ['Closing Net Profit', f"£{net:,.2f}"],
    ]
    return build_pdf_report(
        "Net Profit Report",
        f"Date Range: {date_range}",
        ['Metric', 'Amount'],
        rows,
        [480, 180],
        currency_cols=[1],
    )


@app.route('/reports/export/office-income-pdf')
def export_office_income_pdf():
    """Export office income report as PDF."""
    if not REPORTLAB_AVAILABLE:
        return jsonify({'success': False, 'error': 'PDF library (reportlab) is not installed.'}), 500
    try:
        date_from = request.args.get('date_from') or None
        date_to = request.args.get('date_to') or None
        created_by = request.args.get('created_by') or None
        transactions = db.get_office_transactions(start_date=date_from, end_date=date_to, created_by=created_by)
        income_rows = [t for t in transactions if t.get('transaction_type') in ('Receipt', 'Fee Transfer')]
        total = db.get_office_income_total(date_from, date_to)
        buffer = _build_office_income_pdf(income_rows, date_from, date_to, total)
        fn = f"office_income_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=fn,
            max_age=0,
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reports/export/office-income-csv')
def export_office_income_csv():
    """Export office income report as CSV."""
    date_from = request.args.get('date_from') or None
    date_to = request.args.get('date_to') or None
    created_by = request.args.get('created_by') or None
    transactions = db.get_office_transactions(start_date=date_from, end_date=date_to, created_by=created_by)
    income_rows = [t for t in transactions if t.get('transaction_type') in ('Receipt', 'Fee Transfer')]
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['Date', 'Type', 'Reference', 'Client', 'Amount'])
    for t in income_rows:
        client = f"{t.get('client_code', '')} - {t.get('client_name', '')}" if t.get('client_code') else ''
        writer.writerow([t['transaction_date'], t.get('transaction_type', 'Receipt'), t['reference'], client, t['amount']])
    total = db.get_office_income_total(date_from, date_to)
    writer.writerow(['', '', '', 'Total', str(total)])
    buffer.seek(0)
    fn = f"office_income_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    data = buffer.getvalue().encode('utf-8-sig')
    if request.args.get('save_local'):
        return _save_local_response(data, fn)
    return send_file(io.BytesIO(data), mimetype='text/csv', as_attachment=True, download_name=fn)


@app.route('/reports/export/office-expenses-pdf')
def export_office_expenses_pdf():
    """Export office expenses report as PDF."""
    if not REPORTLAB_AVAILABLE:
        return jsonify({'success': False, 'error': 'PDF library (reportlab) is not installed.'}), 500
    try:
        date_from = request.args.get('date_from') or None
        date_to = request.args.get('date_to') or None
        created_by = request.args.get('created_by') or None
        all_cashbook = db.get_all_cashbook_transactions(start_date=date_from, end_date=date_to, created_by=created_by)
        expenses = [t for t in all_cashbook if t.get('linked_ledger_id') is None and t['transaction_type'] == 'Payment' and t['status'] != 'Declined']
        total = db.get_office_expenses_total(date_from, date_to)
        buffer = _build_office_expenses_pdf(expenses, date_from, date_to, total)
        fn = f"office_expenses_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=fn,
            max_age=0,
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reports/export/office-expenses-csv')
def export_office_expenses_csv():
    """Export office expenses report as CSV."""
    date_from = request.args.get('date_from') or None
    date_to = request.args.get('date_to') or None
    all_cashbook = db.get_all_cashbook_transactions(start_date=date_from, end_date=date_to)
    expenses = [t for t in all_cashbook if t.get('linked_ledger_id') is None and t['transaction_type'] == 'Payment' and t['status'] != 'Declined']
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['Date', 'Reference', 'Source', 'Description', 'Amount', 'Created By'])
    for t in expenses:
        writer.writerow([t['transaction_date'], t['reference'], t['source'], t.get('description') or '', t['amount'], t.get('created_by', 'System')])
    total = db.get_office_expenses_total(date_from, date_to)
    writer.writerow(['', '', '', 'Total', str(total)])
    buffer.seek(0)
    fn = f"office_expenses_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    data = buffer.getvalue().encode('utf-8-sig')
    if request.args.get('save_local'):
        return _save_local_response(data, fn)
    return send_file(io.BytesIO(data), mimetype='text/csv', as_attachment=True, download_name=fn)


@app.route('/reports/export/office-profit-pdf')
def export_office_profit_pdf():
    """Export net profit report as PDF."""
    if not REPORTLAB_AVAILABLE:
        return jsonify({'success': False, 'error': 'PDF library (reportlab) is not installed.'}), 500
    try:
        date_from = request.args.get('date_from') or None
        date_to = request.args.get('date_to') or None
        income_total = db.get_office_income_total(date_from, date_to)
        expenses_total = db.get_office_expenses_total(date_from, date_to)
        buffer = _build_office_profit_pdf(income_total, expenses_total, date_from, date_to)
        fn = f"office_profit_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=fn,
            max_age=0,
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reports/export/office-profit-csv')
def export_office_profit_csv():
    """Export net profit report as CSV."""
    date_from = request.args.get('date_from') or None
    date_to = request.args.get('date_to') or None
    income_total = db.get_office_income_total(date_from, date_to)
    expenses_total = db.get_office_expenses_total(date_from, date_to)
    net = income_total - expenses_total
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['Office Income', str(income_total)])
    writer.writerow(['Office Expenses', str(expenses_total)])
    writer.writerow(['Net Profit', str(net)])
    buffer.seek(0)
    fn = f"office_profit_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    data = buffer.getvalue().encode('utf-8-sig')
    if request.args.get('save_local'):
        return _save_local_response(data, fn)
    return send_file(io.BytesIO(data), mimetype='text/csv', as_attachment=True, download_name=fn)


def _build_xlsx(headers, rows, sheet_title='Report'):
    """Build an Excel workbook in memory and return bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@app.route('/reports/export/ledger-xlsx')
def export_ledger_xlsx():
    """Export client ledger report as Excel."""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'success': False, 'error': 'Excel library (openpyxl) is not installed.'}), 500
    try:
        client_id = request.args.get('client_id', type=int)
        date_from = request.args.get('date_from') or None
        date_to = request.args.get('date_to') or None
        created_by = request.args.get('created_by') or None
        transactions = db.get_ledger_transactions_for_report(client_id, date_from, date_to, created_by)
        headers = ['Transaction ID', 'Date', 'Client Code', 'Client Name', 'Type', 'Reference', 'Source', 'Amount', 'Created By']
        rows = []
        for t in transactions:
            rows.append([t.get('transaction_id', ''), t['transaction_date'], t.get('client_code', ''), t.get('client_name', ''),
                         t['transaction_type'], t['reference'], t['source'], float(t['amount']), t.get('created_by', 'System')])
        data = _build_xlsx(headers, rows, 'Client Ledger')
        fn = f"ledger_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return _xlsx_download_response(data, fn)
    except Exception as e:
        logger.exception('Excel export failed: ledger')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reports/export/cashbook-xlsx')
def export_cashbook_xlsx():
    """Export cashbook report as Excel."""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'success': False, 'error': 'Excel library (openpyxl) is not installed.'}), 500
    try:
        date_from = request.args.get('date_from') or None
        date_to = request.args.get('date_to') or None
        created_by = request.args.get('created_by') or None
        transactions = db.get_all_cashbook_transactions(start_date=date_from, end_date=date_to, created_by=created_by)
        headers = ['Transaction ID', 'Date', 'Type', 'Reference', 'Source', 'Status', 'Client', 'Amount', 'Created By']
        rows = []
        for t in transactions:
            client = f"{t.get('client_code', '')} - {t.get('client_name', '')}".strip(' -') or 'Standalone'
            rows.append([t.get('transaction_id', ''), t['transaction_date'], t['transaction_type'], t['reference'],
                         t['source'], t['status'], client, float(t['amount']), t.get('created_by', 'System')])
        data = _build_xlsx(headers, rows, 'Cashbook')
        fn = f"cashbook_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return _xlsx_download_response(data, fn)
    except Exception as e:
        logger.exception('Excel export failed: cashbook')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reports/export/office-income-xlsx')
def export_office_income_xlsx():
    """Export office income report as Excel."""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'success': False, 'error': 'Excel library (openpyxl) is not installed.'}), 500
    try:
        date_from = request.args.get('date_from') or None
        date_to = request.args.get('date_to') or None
        created_by = request.args.get('created_by') or None
        transactions = db.get_office_transactions(start_date=date_from, end_date=date_to, created_by=created_by)
        income_rows = [t for t in transactions if t.get('transaction_type') in ('Receipt', 'Fee Transfer')]
        total = db.get_office_income_total(date_from, date_to)
        headers = ['Date', 'Type', 'Reference', 'Client', 'Amount']
        rows = []
        for t in income_rows:
            client = f"{t.get('client_code', '')} - {t.get('client_name', '')}" if t.get('client_code') else ''
            rows.append([t['transaction_date'], t.get('transaction_type', 'Receipt'), t['reference'], client, float(t['amount'])])
        rows.append(['', '', '', 'Total', float(total)])
        data = _build_xlsx(headers, rows, 'Office Income')
        fn = f"office_income_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return _xlsx_download_response(data, fn)
    except Exception as e:
        logger.exception('Excel export failed: office income')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reports/export/office-expenses-xlsx')
def export_office_expenses_xlsx():
    """Export office expenses report as Excel."""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'success': False, 'error': 'Excel library (openpyxl) is not installed.'}), 500
    try:
        date_from = request.args.get('date_from') or None
        date_to = request.args.get('date_to') or None
        all_cashbook = db.get_all_cashbook_transactions(start_date=date_from, end_date=date_to)
        expenses = [t for t in all_cashbook if t.get('linked_ledger_id') is None and t['transaction_type'] == 'Payment' and t['status'] != 'Declined']
        total = db.get_office_expenses_total(date_from, date_to)
        headers = ['Date', 'Reference', 'Source', 'Description', 'Amount', 'Created By']
        rows = []
        for t in expenses:
            rows.append([t['transaction_date'], t['reference'], t['source'], t.get('description') or '', float(t['amount']), t.get('created_by', 'System')])
        rows.append(['', '', '', 'Total', float(total), ''])
        data = _build_xlsx(headers, rows, 'Office Expenses')
        fn = f"office_expenses_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return _xlsx_download_response(data, fn)
    except Exception as e:
        logger.exception('Excel export failed: office expenses')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/reports/export/office-profit-xlsx')
def export_office_profit_xlsx():
    """Export net profit report as Excel."""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'success': False, 'error': 'Excel library (openpyxl) is not installed.'}), 500
    try:
        date_from = request.args.get('date_from') or None
        date_to = request.args.get('date_to') or None
        income_total = db.get_office_income_total(date_from, date_to)
        expenses_total = db.get_office_expenses_total(date_from, date_to)
        net = income_total - expenses_total
        headers = ['Category', 'Amount']
        rows = [['Office Income', float(income_total)], ['Office Expenses', float(expenses_total)], ['Net Profit', float(net)]]
        data = _build_xlsx(headers, rows, 'Net Profit')
        fn = f"office_profit_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return _xlsx_download_response(data, fn)
    except Exception as e:
        logger.exception('Excel export failed: office profit')
        return jsonify({'success': False, 'error': str(e)}), 500


def _get_exports_dir() -> str:
    """Return writable exports directory for save_local / desktop exports."""
    exports = os.path.join(_get_data_dir(), 'exports')
    os.makedirs(exports, exist_ok=True)
    return exports


def _xlsx_download_response(data: bytes, filename: str):
    """Return Excel download from memory; write to disk only for save_local."""
    if request.args.get('save_local'):
        return _save_local_response(data, filename)
    return send_file(
        io.BytesIO(data),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def _save_local_response(file_bytes: bytes, filename: str):
    """Save file to local exports folder and return JSON response."""
    exports_dir = _get_exports_dir()
    filepath = os.path.join(exports_dir, filename)
    with open(filepath, 'wb') as f:
        f.write(file_bytes)
    return jsonify({
        'success': True,
        'message': 'Export successful.',
        'filename': filename,
        'filepath': filepath,
        'folder': exports_dir,
    })


@app.route('/reports/export/open-folder')
def open_exports_folder():
    """Open the local exports folder in Windows Explorer."""
    exports_dir = _get_exports_dir()
    try:
        import subprocess
        subprocess.Popen(['explorer', exports_dir])
    except Exception:
        pass
    flash('Exports folder opened.', 'info')
    return redirect(url_for('reports'))


@app.route('/reports/export/open-file')
def open_export_file():
    """Open a specific exported file."""
    filepath = request.args.get('path', '')
    if filepath and os.path.isfile(filepath):
        try:
            os.startfile(filepath)
        except Exception:
            pass
    return ('', 204)


if __name__ == '__main__':
    # Create templates and static directories if they don't exist
    os.makedirs('templates', exist_ok=True)
    os.makedirs('static', exist_ok=True)
    
    print("\n" + "="*70)
    print("Nexal Legal — Client Ledger & Compliance System")
    print("="*70)
    print(f"\nStarting server...")
    print(f"Open your browser and navigate to: http://127.0.0.1:5001")
    print(f"\nPress CTRL+C to stop the server\n")
    print("="*60 + "\n")
    
    app.run(debug=True, host='127.0.0.1', port=5001)
